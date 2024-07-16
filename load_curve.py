import pandas as pd
import geopandas as gpd
from shapely.geometry import Point
import re
import urllib.request
from io import StringIO
from io import BytesIO
from zipfile import ZipFile
import demandlib.bdew as bdew
import datetime
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import subprocess
from openpyxl.drawing.image import Image

class Temperature:
    def __init__(self,url):
        self.url = url
        self.url_all = url + 'TU_Stundenwerte_Beschreibung_Stationen.txt'
  
    def stationsfromtxt(self):
        # open file from url
        # replace the whitespaces by ; - take attention to column "Stationsname" as these names may be separated by whitespace but should not match to different columns
        filestring = ''
        for line in urllib.request.urlopen(self.url):
            s = line.decode('latin1')
            if s[0] == 'S':
                # this is the first line
                s = re.sub("\s+", ";", s.strip())
            elif s[0] == "-":
                continue
            else:
                s = re.sub("([0-9])(\s+)([0-9])", r'\1;\3', s.strip())
                s = re.sub("\s{2,}", r';', s)
                s = re.sub("([0-9])(\s+)", r'\1;', s)
            filestring += s + '\n'
        output = StringIO(filestring)
        
        allstationdf = pd.read_csv(
            output, 
            delimiter=";",
            dtype={
            'Stations_id': 'string',
            'Stationsname': 'string',
            'Bundesland': 'string',
            })

        geo_df = gpd.GeoDataFrame(allstationdf, crs=4326, geometry=[Point(xy) for xy in zip(allstationdf.geoLaenge, allstationdf.geoBreite)])
        geo_df = geo_df.to_crs(25832)

        return geo_df

    def nearestStation(self, poi, gdf, year, i=10):
        '''
        searches gdf for number of nearest stations to point(x,y) - x,y in EPSG:25832 / UTM32N / ETRS89
        poi: Point of interest. Place where temp data is needed
        gdf: gdf with all stations
        year: year for which temperature data is wanted
        '''
        x = poi[0]
        y = poi[1]
        # explizite Kopie des GeoDataFrame
        gdf_copy = gdf.copy()

        # Entfernungen hinzufügen
        gdf_copy['distance'] = [Point(x, y).distance(gdf_copy['geometry'][i]) for i in gdf_copy.index]
        alt = gdf_copy.nsmallest(5, 'distance').reset_index(drop=True)
        try:

            # Filtere Stationen, die das Jahr enthalten
            gdf_copy = gdf_copy[(gdf_copy['von_datum'] < (year-i+1) * 10000) & (gdf_copy['bis_datum'] > (year + 1) * 10000)]

            # kleinste Entfernung
            ns = gdf_copy.nsmallest(1, 'distance').reset_index(drop=True)
            return ns
        except:
            print('Keine Station mit den gewählten Parameter gefunden. \nAlternatieven: ')
            print(alt)

    def tempdata(self, url, station_id, year, start_date, end_date, n =10):
        '''loads and returns mean temperature data from last 5 years as Dataframe'''
        zipfile = f'stundenwerte_TU_{station_id}_{start_date}_{end_date}_hist.zip'
        file = f'produkt_tu_stunde_{start_date}_{end_date}_{station_id}.txt'
        url = urllib.request.urlopen(url+zipfile)

        with ZipFile(BytesIO(url.read())) as my_zip_file:
            data = pd.read_csv(
                my_zip_file.open(file), 
                delimiter=';',
                skipinitialspace=True
            )
        # mean
        dataframes = []
        for i in range(n):
            filtered_data = data[data['MESS_DATUM'].astype(str).str.startswith(str(year-i))]['TT_TU'].reset_index(drop=True)
            if len(filtered_data) == 8760:
                dataframes.append(filtered_data)
            else: 
                filtered_data = filtered_data.drop(filtered_data.index[1416:1440]).reset_index(drop=True)
                dataframes.append(filtered_data)
        average_data = pd.concat(dataframes).groupby(level=0).mean().reset_index(drop=True)
        return average_data
    
class EnergyDemandProfile:
    """
    Class for simulating standard load profiles for heta and electricity based on BDEW.

    Parameters
    ----------
    year : int
        ...
    """
    def __init__(self, year, temperature_data, holidays):
        self.year = year
        self.temperature = temperature_data
        self.holidays = holidays
        self.demand_time_series = pd.date_range(start=datetime.datetime(year, 1, 1, 0),
                                end=datetime.datetime(year, 12, 31, 23),
                                freq='H')

    def create_heat_demand_profile(self, building_type, building_class, wind_class, ww_incl, annual_heat_demand):
        # Baualtersklasse 1-11. NRW:3 Quelle:Praxisinformation P 2006 / 8 Gastransport / Betriebswirtschaft, BGW, 2006, Seite 43 Tabelle 2 und 3 
        heat_demand = bdew.HeatBuilding(self.demand_time_series, holidays=self.holidays, temperature=self.temperature,
                                      shlp_type=building_type, building_class=building_class, 
                                      wind_class=wind_class, ww_incl=ww_incl, annual_heat_demand=annual_heat_demand, 
                                      name=building_type).get_bdew_profile()
        return heat_demand

    def create_power_demand_profile(self, annual_electricity_demand_per_sector):
        elec_slp = bdew.ElecSlp(self.year, holidays=self.holidays)
        # Further logic
        elec_demand = elec_slp.get_profile(annual_electricity_demand_per_sector)
        return elec_demand

    def plot_bar_chart(self, dataframe, column_names, figsize=(18, 4), colors=['blue', 'orange'], filename='../Lastprofil.png'):
        plt.figure()  # new figure
        for i in range(len(column_names)):
            plt.bar(range(len(dataframe)), dataframe[column_names[i]], color=colors[i], label=column_names[i], width =1)
            
        # Plot size
        fig = plt.gcf()
        fig.set_size_inches(figsize)

        plt.xlabel('Zeit [h]')
        plt.ylabel('Wärmebedarf und Verlust [MW]')
        plt.title('Wärmebedarf und Verlust pro Stunde im Jahr')
        plt.legend()
        plt.savefig(filename, bbox_inches='tight')
        #plt.show()
        plt.close(fig)

    def set_up_df(self,year,resolution,freq):
        '''Creates a DataFrame for collecting generated profiles'''
        demand = pd.DataFrame(
            index=pd.date_range(datetime.datetime(year, 1, 1, 0), periods=resolution, freq=freq))
        return demand
    
class LoadProfile:
    '''Class to manage methods for the excel file, where the net report is stored'''
    def __init__(self, net_result, excel_path):
        self.net_result = net_result
        self.path = excel_path

    # def import_sheet(self,sheet=0):
    #     '''import excel sheet data as pandas data frame'''
    #     self.imported_df = pd.read_excel(self.path,sheet_name=sheet)
    @staticmethod
    def sort_columns_by_sum(df):
        """
        arranges columns of a data frame (df) in ascending order.
        returns sorted data frame
        """
        # Spalten nach Summe sortieren und DataFrame neu anordnen
        sorted_columns = df.sum().sort_values().index
        sorted_df = df[sorted_columns]
        return sorted_df
    @staticmethod
    def add_loss(demand_df, df, resolution = 8760):
        loss_sum = df['Verlust [MWh/a]'].max()
        loss_hourly = loss_sum/resolution
        demand_df['Verlust'] = loss_hourly
        return demand_df
    @staticmethod
    def add_sum_buildings(df):
        '''Adds a sum column to a copy of a data frame and returns it'''
        df_with_sum = df.copy()
        df_with_sum['Summe aller Gebäudetypen'] = df.sum(axis=1)
        return df_with_sum
    @staticmethod
    def add_sum(df):
        df_sum = df.copy()
        df_sum['Gesamtsumme'] = df_sum['Summe aller Gebäudetypen']+df_sum['Verlust']
        return df_sum

    def safe_in_excel(self, df, col = 0, index_bool=False, sheet_option ='replace', sheet = 'Lastprofil'):
        '''
        saves data frame in the stated excel file
        df: data frame
        col: startcolumn
        index_bool: specify if df index is added in excel
        sheet_option: specifies the writing option for ExcelWriter--> (parameter) if_sheet_exists: Literal['error', 'new', 'replace', 'overlay']
        sheet: desired excel sheet
        '''
        filename = self.path
        # open excel file and write the data frame in stated sheet
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists=sheet_option) as writer:
            df.to_excel(writer, sheet_name=sheet, index=index_bool, startcol=col)

        # adjust cell size
        wb = load_workbook(filename = filename)        
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length+1)
            ws.column_dimensions[column].width = adjusted_width
        wb.save(filename)

    def embed_image_in_excel(self, row, col, sheet = 'Lastprofil', image_filename = '../Lastprofil.png'):
        filename = self.path
        # Laden Sie die vorhandene Excel-Datei
        workbook = load_workbook(filename)

        # Holen Sie sich das Arbeitsblatt
        worksheet = workbook[sheet]

        # Fügen Sie das Bild in das Excel-Sheet ein
        img = Image(image_filename)
        worksheet.add_image(img, f'{chr(65 + col)}{row + 1}')

        # Speichern Sie die Excel-Datei
        workbook.save(filename)

    def open_excel_file(self):
        try:
            subprocess.Popen(['start', 'excel', self.path], shell=True)
        except Exception as e:
            print(f"Fehler beim Öffnen der Excel-Datei: {e}")
