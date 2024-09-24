import pandas as pd
import geopandas as gpd
from shapely.geometry import Point
import re
import urllib.request
from io import StringIO
from io import BytesIO
from zipfile import ZipFile
import demandlib.bdew as bdew
import datetime
import matplotlib.pyplot as plt
from openpyxl import load_workbook
import subprocess
from openpyxl.drawing.image import Image


class Temperature:
    '''
    A class to handle temperature data from a specified URL.

    Attributes
    ----------
    url : str
        Base URL for temperature data.
    url_all : str
        URL for the file containing station descriptions.

    Methods
    -------
    stationsfromtxt():
        Retrieves and processes station data from the URL.
    nearestStation(poi, gdf, year, i=10):
        Finds the nearest station to a given point of interest (POI) for a specific year.
    tempdata(url, station_id, year, start_date, end_date, n=10):
        Loads and returns mean temperature data from the last n years as a DataFrame.
    '''

    def __init__(self,url):
        '''
        Initializes the Temperature class with the base URL for temperature data.

        Parameters
        ----------
        url : str
            Base URL for temperature data.
        '''
        self.url = url
        self.url_all = url + 'TU_Stundenwerte_Beschreibung_Stationen.txt'
  
    def stationsfromtxt(self):
        '''
        Retrieves and processes weather station data from the URL.

        Returns
        -------
        GeoDataFrame
            A GeoDataFrame with station data including geometry for spatial operations.
        '''
        # open file from url
        # replace the whitespaces by ; - take attention to column "Stationsname" as these names may be separated by whitespace but should not match to different columns
        filestring = ''
        for line in urllib.request.urlopen(self.url_all):
            s = line.decode('latin1')
            if s[0] == 'S':
                # this is the first line
                s = re.sub("\s+", ";", s.strip())
            elif s[0] == "-":
                continue
            else:
                s = re.sub("([0-9])(\s+)([0-9])", r'\1;\3', s.strip())
                s = re.sub("\s{2,}", r';', s)
                s = re.sub("([0-9])(\s+)", r'\1;', s)
            filestring += s + '\n'
        output = StringIO(filestring)
        
        allstationdf = pd.read_csv(
            output, 
            delimiter=";",
            dtype={
            'Stations_id': 'string',
            'von_datum': 'string',
            'bis_datum': 'string',
            'Stationshoehe': 'float',
            'geoBreite': 'float',
            'geoLaenge': 'float',
            'Stationsname': 'string',
            'Bundesland': 'string',
            'Abgabe': 'string'
            })

        geo_df = gpd.GeoDataFrame(allstationdf, crs=4326, geometry=[Point(xy) for xy in zip(allstationdf.geoLaenge, allstationdf.geoBreite)])
        geo_df = geo_df.to_crs(25832)

        return geo_df

    def nearestStation(self, poi, gdf, i=10):
        '''
        Searches for the nearest stations to a point of interest (POI) for a specific year.

        Parameters
        ----------
        poi : tuple
            Point of interest (x, y) where temperature data is needed.
        gdf : GeoDataFrame
            GeoDataFrame with all stations.
        i : int, optional
            Number of nearest stations to consider (default is 10).

        Returns
        -------
        GeoDataFrame
            A GeoDataFrame with the nearest station(s).
        '''
        current_year = datetime.datetime.now().year
        x = poi[0]
        y = poi[1]
        gdf_copy = gdf.copy()

        # Add distances
        gdf_copy['distance'] = [Point(x, y).distance(gdf_copy['geometry'][i]) for i in gdf_copy.index]
        try:

            # Filter stations that have current temperature data available and go back i years
            gdf_filter = gdf_copy[(gdf_copy['von_datum'] < (current_year-(i+1)) * 10000) & (gdf_copy['bis_datum'] > (current_year-1) * 10000)] # historical data goes up to 20231231 for year 2024

            # Smallest distance
            ns = gdf_filter.nsmallest(1, 'distance').reset_index(drop=True)
            return ns
        except:
            print('No station found with current temperature data. Going for closest older Station')
            try: 
                ns = gdf_copy.nsmallest(1, 'distance').reset_index(drop=True)
                return ns
            except Exception as e:
                print('Error: '+e)



    def tempdata(self, url, station_id, start_date, end_date, n = 10):
        '''
        Loads and returns mean temperature data from the last n years as a DataFrame.

        Parameters
        ----------
        url : str
            Base URL for temperature data.
        station_id : str
            ID of the station.
        year : int
            Year for which temperature data is wanted.
        start_date : str
            Start date for the data.
        end_date : str
            End date for the data.
        n : int, optional
            Number of years to consider (default is 10).

        Returns
        -------
        DataFrame
            DataFrame containing the mean temperature data.
        '''
        if end_date[-4:] != '1231':
            endyear = int(end_date[:4])-1
            end_date = endyear*10000+1231
        else:
            endyear = end_date[:4]
        zipfile = f'stundenwerte_TU_{station_id}_{start_date}_{end_date}_hist.zip'
        file = f'produkt_tu_stunde_{start_date}_{end_date}_{station_id}.txt'
        url = urllib.request.urlopen(url+zipfile)

        with ZipFile(BytesIO(url.read())) as my_zip_file:
            data = pd.read_csv(
                my_zip_file.open(file), 
                delimiter=';',
                skipinitialspace=True
            )
        # Mean
        dataframes = []
        for i in range(n):
            filtered_data = data[data['MESS_DATUM'].astype(str).str.startswith(str(endyear-i))]['TT_TU'].reset_index(drop=True)
            if len(filtered_data) == 8760:
                dataframes.append(filtered_data)
            else: 
                filtered_data = filtered_data.drop(filtered_data.index[1416:1440]).reset_index(drop=True)
                dataframes.append(filtered_data)
        average_data = pd.concat(dataframes).groupby(level=0).mean().reset_index(drop=True)
        return average_data
    
def safe_in_excel(path, df, col = 0, index_bool=False, sheet = 'Lastprofil'):
    '''
    Saves the dataframe to the specified Excel file and sheet.

    Parameters
    ----------
    df : pd.DataFrame
        The dataframe to save.
    col : int, optional
        The starting column index in the Excel sheet (default is 0).
    index_bool : bool, optional
        Whether to include the dataframe index in the Excel file (default is False).
    sheet_option : str, optional
        The option for handling existing sheets (default is 'replace'). Options: 'error', 'new', 'replace', 'overlay'.
    sheet : str, optional
        The name of the Excel sheet (default is 'Lastprofil').
    '''
    filename = path
    # Open excel file and write the data frame in stated sheet
    with pd.ExcelWriter(filename, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, sheet_name=sheet, index=index_bool, startcol=col)

    # Adjust cell size
    wb = load_workbook(filename = path)        
    ws = wb[sheet]
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length+1)
        ws.column_dimensions[column].width = adjusted_width
    wb.save(filename)

class LoadProfile:
    '''
     A class for managing and analyzing load profiles, including generating, sorting, and saving energy demand profiles.

    Attributes
    ----------
    net_result : object
        The result object containing network data.
    path : str
        The path to the Excel file where results are stored.
    year : int
        The year for which the demand profiles are generated.
    temperature : pd.Series
        The temperature data series for the year.
    holidays : list
        List of holidays in the year.
    demand_time_series : pd.DatetimeIndex
        Time series index for the entire year at hourly frequency.

    Methods
    -------
    create_heat_demand_profile(building_type, building_class, wind_class, ww_incl, annual_heat_demand):
        Creates a heating demand profile based on building characteristics and annual heat demand.
    set_up_df(year, resolution, freq):
        Creates a DataFrame for collecting generated profiles with the specified resolution and frequency.
    sort_columns_by_sum(df):
        Sorts the columns of a dataframe in ascending order based on their sum and returns the sorted dataframe.
    add_loss(demand_df, df, resolution=8760):
        Adds a loss column to the demand dataframe based on the maximum annual loss in another dataframe.
    add_sum_buildings(df):
        Adds a column for the sum of all building types in a dataframe and returns the modified dataframe.
    add_sum(df):
        Adds a total sum column that includes the sum of all building types and losses.
    plot_bar_chart(dataframe, column_names, figsize=(18, 4), colors=['blue', 'orange'], filename='../Lastprofil.png'):
        Plots a bar chart of specified columns in a dataframe.
    save_in_excel(df, col=0, index_bool=False, sheet_option='replace', sheet='Lastprofil'):
        Saves the dataframe to the specified Excel file and sheet.
    embed_image_in_excel(row, col, sheet='Lastprofil', image_filename='../Lastprofil.png'):
        Embeds an image into the specified Excel sheet at a given position.
    open_excel_file():
        Opens the Excel file using the default application.
    '''

    def __init__(self, net_result, excel_path, year, temperature_data, holidays):
        '''
        Initializes the LoadProfile class with the network result object, path to the Excel file, year, temperature data, and holidays.

        Parameters
        ----------
        net_result : object
            The result object containing network data.
        excel_path : str
            The path to the Excel file where results are stored.
        year : int
            The year for which the demand profiles are generated.
        temperature_data : pd.Series
            The temperature data series for the year.
        holidays : list
            List of holidays in the year.
        '''
        self.net_result = net_result
        self.path = excel_path
        self.year = year
        self.temperature = temperature_data
        self.holidays = holidays
        self.demand_time_series = pd.date_range(start=datetime.datetime(year, 1, 1, 0),
                                end=datetime.datetime(year, 12, 31, 23),
                                freq='H')
    
    def create_heat_demand_profile(self, building_type, building_class, wind_class, ww_incl, annual_heat_demand):
        '''
        Creates a heating demand profile based on building characteristics and annual heat demand.

        Parameters
        ----------
        building_type : str
            The type of the building (e.g. EFH = Einfamilienhaus).
        building_class : int
            The building age class (1-11).
        wind_class : int
            The wind load class.
        ww_incl : bool
            Whether domestic hot water (DHW) is included.
        annual_heat_demand : float
            The annual heat demand in MWh.

        Returns
        -------
        pd.Series
            The generated heat demand profile.
        '''
        # Building age class 1-11. NRW=3, source: Praxisinformation P 2006 / 8 Gastransport / Betriebswirtschaft, BGW, 2006, Seite 43 Tabelle 2 und 3 
        heat_demand = bdew.HeatBuilding(self.demand_time_series, holidays=self.holidays, temperature=self.temperature,
                                      shlp_type=building_type, building_class=building_class, 
                                      wind_class=wind_class, ww_incl=ww_incl, annual_heat_demand=annual_heat_demand, 
                                      name=building_type).get_bdew_profile()
        return heat_demand

    @staticmethod
    def set_up_df(year,resolution,freq):
        '''
        Creates a DataFrame for collecting generated profiles with the specified resolution and frequency.

        Parameters
        ----------
        year : int
            The year for the time series index.
        resolution : int
            The number of periods in the time series.
        freq : str
            Frequency of the time series (e.g., 'H' for hourly).

        Returns
        -------
        pd.DataFrame
            DataFrame with a time series index.
        '''
        demand = pd.DataFrame(
            index=pd.date_range(datetime.datetime(year, 1, 1, 0), periods=resolution, freq=freq))
        return demand

    @staticmethod
    def sort_columns_by_sum(df):
        '''Sorts the columns of a dataframe in ascending order based on their sum.

        Parameters
        ----------
        df : pd.DataFrame
            The dataframe whose columns need to be sorted.

        Returns
        -------
        pd.DataFrame
            Dataframe with columns sorted by their sum.
        '''
    
        sorted_columns = df.sum().sort_values().index
        sorted_df = df[sorted_columns]
        return sorted_df
    
    @staticmethod
    def add_loss(demand_df, df, resolution = 8760):
        '''
        Adds a loss column to the demand dataframe based on the maximum annual loss in another dataframe.

        Parameters
        ----------
        demand_df : pd.DataFrame
            The dataframe to which the loss column will be added.
        df : pd.DataFrame
            The dataframe containing loss information.
        resolution : int, optional
            The time resolution in hours (default is 8760 for an hourly resolution over a year).

        Returns
        -------
        pd.DataFrame
            The modified demand dataframe with the added loss column.
        '''
        loss_sum = df['Verlust [MWh/a]'].max()
        loss_extra_sum = df['Verlust bei extra Daemmung [MWh/a]'].max() # extra insulation
        loss_hourly = loss_sum/resolution
        loss_extra_hourly = loss_extra_sum/resolution
        demand_df['Verlust'] = loss_hourly
        demand_df['Verlust bei extra Dämmung'] = loss_extra_hourly
        return demand_df
    
    @staticmethod
    def add_sum_buildings(df):
        '''
        Adds a column for the sum of all building types in a dataframe.

        Parameters
        ----------
        df : pd.DataFrame
            The dataframe to which the sum column will be added.

        Returns
        -------
        pd.DataFrame
            The modified dataframe with the sum column.
        '''
        df_with_sum = df.copy()
        df_with_sum['Summe aller Gebäudetypen'] = df.sum(axis=1)
        return df_with_sum
    
    # @staticmethod
    # def add_glf(df,glf):
    #     glf = 1
    #     df_glf = df.copy()

    #     # Calculate the average load
    #     avg_load = df_glf['Summe aller Gebäudetypen'].mean()

    #     # Apply the formula to adjust the load profile with the coincidence factor
    #     df_glf['Summe aller Gebäudetypen mit GLF'] = avg_load + (df_glf['Summe aller Gebäudetypen'] - avg_load) * glf

    #     return df_glf
    
    @staticmethod
    def add_sum(df):
        '''
        Adds a total sum column that includes the sum of all building types and losses.

        Parameters
        ----------
        df : pd.DataFrame
            The dataframe to which the total sum column will be added.

        Returns
        -------
        pd.DataFrame
            The modified dataframe with the total sum column.
        '''
        df_sum = df.copy()
        # df_sum['Gesamtsumme'] = df_sum['Summe aller Gebäudetypen mit GLF']+df_sum['Verlust']
        df_sum['Gesamtsumme'] = df_sum['Summe aller Gebäudetypen']+df_sum['Verlust']
        df_sum['Gesamtsumme (extra Dämmung)'] = df_sum['Summe aller Gebäudetypen']+df_sum['Verlust bei extra Dämmung']
        return df_sum
    
    @staticmethod
    def plot_bar_chart(dataframe, column_names, figsize=(18, 4), colors=['blue', 'orange'], filename='../Lastprofil.png', ylabel='Wärmebedarf und Verlust [MW]', title='Wärmebedarf und Verlust pro Stunde im Jahr'):
        '''
        Plots a bar chart of specified columns in a dataframe.

        Parameters
        ----------
        dataframe : pd.DataFrame
            The data frame containing the data to plot.
        column_names : list
            List of column names to plot.
        figsize : tuple, optional
            Size of the figure (default is (18, 4)).
        colors : list, optional
            List of colors for the bars (default is ['blue', 'orange']).
        filename : str, optional
            The filename to save the plot (default is '../Lastprofil.png').
        '''
        plt.figure()  # new figure
        for i in range(len(column_names)):
            plt.bar(range(len(dataframe)), dataframe[column_names[i]], color=colors[i], label=column_names[i], width =1)
            
        # Plot size
        fig = plt.gcf()
        fig.set_size_inches(figsize)

        plt.xlabel('Zeit [h]')
        plt.ylabel(ylabel)
        plt.title(title)
        plt.legend()
        plt.savefig(filename, bbox_inches='tight')
        plt.close(fig)

    def save_in_excel(self, df, col = 0, index_bool=False, sheet_option ='replace', sheet = 'Lastprofil'):
        '''
        Saves the dataframe to the specified Excel file and sheet.

        Parameters
        ----------
        df : pd.DataFrame
            The dataframe to save.
        col : int, optional
            The starting column index in the Excel sheet (default is 0).
        index_bool : bool, optional
            Whether to include the dataframe index in the Excel file (default is False).
        sheet_option : str, optional
            The option for handling existing sheets (default is 'replace'). Options: 'error', 'new', 'replace', 'overlay'.
        sheet : str, optional
            The name of the Excel sheet (default is 'Lastprofil').
        '''
        filename = self.path
        # Open excel file and write the data frame in stated sheet
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists=sheet_option) as writer:
            df.to_excel(writer, sheet_name=sheet, index=index_bool, startcol=col)

        # Adjust cell size
        wb = load_workbook(filename = filename)        
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter # Get the column name
            for cell in col:
                try: # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length+1)
            ws.column_dimensions[column].width = adjusted_width
        wb.save(filename)

    def embed_image_in_excel(self, row, col, sheet = 'Lastprofil', image_filename = '../Lastprofil.png'):
        '''
        Embeds an image into the specified Excel sheet at a given position.

        Parameters
        ----------
        row : int
            The row index where the image should be placed.
        col : int
            The column index where the image should be placed.
        sheet : str, optional
            The name of the Excel sheet (default is 'Lastprofil').
        image_filename : str, optional
            The filename of the image to embed (default is '../Lastprofil.png').
        '''
        filename = self.path
        workbook = load_workbook(filename)
        worksheet = workbook[sheet]
        img = Image(image_filename)
        worksheet.add_image(img, f'{chr(65 + col)}{row + 1}')
        workbook.save(filename)

    def open_excel_file(self):
        '''
        Opens the Excel file using the default application.
        '''
        try:
            subprocess.Popen(['start', 'excel', self.path], shell=True)
        except Exception as e:
            print(f"Fehler beim Öffnen der Excel-Datei: {e}")
